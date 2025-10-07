from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def update_report_tables(ws, table_data_dict, start_cell, gap, titles=None, formula_columns=None):
    """
    Update Excel tables in the worksheet with new data, optional section titles, and optional formulas.

    Args:
        ws: The worksheet object.
        table_data_dict: dict of {table_name: DataFrame}.
        start_cell: Cell string like 'B10' indicating where to begin writing.
        gap: Number of blank rows between tables.
        titles: Optional dict of {table_name: title string}.
        formula_columns: Optional dict of {table_name: {column_name: formula_string}}.
                        Formulas must use structured references if desired.
    """

    start_col_letter = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    start_col =   ws[start_cell].column

    current_row = start_row

    for table_name, df in table_data_dict.items():
        if table_name not in ws.tables:
            raise ValueError(f"Table '{table_name}' not found in worksheet.")

        table = ws.tables[table_name]

        excel_columns = [col.name for col in table.tableColumns]

        total_cols = len(excel_columns)
        nrows = df.shape[0]

        col_name_to_index = {name: idx for idx, name in enumerate(excel_columns)}

        # Step 0: Write title if available
        if titles and table_name in titles:
            ws.cell(row=current_row, column=start_col, value=titles[table_name])
            current_row += 1

        # Step 1: Write header row
        for col_name, col_offset in col_name_to_index.items():
            ws.cell(row=current_row, column=start_col + col_offset, value=col_name)

        start_data_row = current_row + 1

        # Step 2: Write data rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_data_row):
            for df_col_idx, col_name in enumerate(df.columns):
                if col_name in col_name_to_index:
                    col_offset = col_name_to_index[col_name]
                    ws.cell(row=r_idx, column=start_col + col_offset, value=row[df_col_idx])

        # Step 3: Resize table range
        end_col = start_col + total_cols - 1
        end_row = start_data_row + nrows - 1
        table.ref = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{end_row}"

        # Step 4: Apply formulas if applicable
        if formula_columns and table_name in formula_columns:
            for col_name, formula_template in formula_columns[table_name].items():
                if col_name not in col_name_to_index:
                    raise ValueError(f"Formula column '{col_name}' not found in table '{table_name}'")

                formula_col_offset = col_name_to_index[col_name]
                cusip_col_offset = col_name_to_index["Cusip"]
                print(f"[DEBUG] Applying formula to table '{table_name}', column '{col_name}'")
                print(f"[DEBUG] Formula content: {formula_template}")
                for r_idx in range(start_data_row, start_data_row + nrows):
                    cusip_cell = ws.cell(row=r_idx, column=start_col + cusip_col_offset).coordinate
                    # Replace placeholder in formula_template with actual cell ref
                    formula = formula_template.replace("{cusip_cell}", cusip_cell)
                    ws.cell(row=r_idx, column=start_col + formula_col_offset).value = formula

        # Debug print
        print(f"Updated table: {table_name}, ref: {table.ref}")

        # Step 5: Advance to next block
        current_row = end_row + gap + 1

    return
