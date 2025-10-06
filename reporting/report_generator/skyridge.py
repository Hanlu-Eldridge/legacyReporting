import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime, date
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# palceholder for datasets, to be replaced with Snowflake holdings table, transaction table and security master table
#  . .\.venv\Scripts\Activate.ps1
# dir = "C:/Users/XiaHanlu/workspace/legacyReporting/local_reading/panagram"
# date = '04-30-2025'
def process_raw_data(date: str, inputdir: str):
    #change this if using different portfolio => move to config
    choices_port = ['44G8/813/IG CLO FWH', '44G9/814/ABS FWH', '44H1/815/Atypical FWH', '44I4/816/HY CLOs FWH', 'P44I5/817/HY CLOs OC', '44J6/821/IG CLO OC',
                   '44J7/822/ABS OC', '44J5/823/Atypical OC']
    target_portfolio_list = ['SBL_813_813','SBL_814_814','SBL_815_815','SBL_816_816','SBL_817_817','SBL_821_821','SBL_822_822','SBL_823_823']

    df_holding = pd.read_excel(inputdir + '/Panagram Holding File ' + date + '.xlsx',
                               sheet_name='DBO_PANAGRAM_SBLHoldingFile_NoF', index_col=None, header=0)
    df_holding.columns = df_holding.columns.str.strip()
    df_holding = df_holding[df_holding['Portfolio Code'].isin(target_portfolio_list)]
    df_trader_info = pd.read_excel(inputdir + '/Input_From_Jared.xlsx', sheet_name='Sheet1', index_col=None, header=0)


    # creating transactions table
    df_transaction = pd.read_excel(inputdir + '/Security Transactions.xlsx', sheet_name='DBO_Security Transactions',
                                   index_col=None, header=0,
                                   usecols=['Portfolio', 'Security ID','Tran Type', 'Security Description',
                                            'Trade Date', 'Settle Date', 'Maturity Date', 'Quantity', 'Coupon Rate',
                                            'Price', 'Cost Proceeds'])
    df_transaction = df_transaction[df_transaction['Portfolio'].isin(target_portfolio_list)]
    conditions_port = [
        df_transaction['Portfolio'] == 'SBL_813_813',
        df_transaction['Portfolio'] == 'SBL_814_814',
        df_transaction['Portfolio'] == 'SBL_815_815',
        df_transaction['Portfolio'] == 'SBL_816_816',
        df_transaction['Portfolio'] == 'SBL_817_817',
        df_transaction['Portfolio'] == 'SBL_821_821',
        df_transaction['Portfolio'] == 'SBL_822_822',
        df_transaction['Portfolio'] == 'SBL_823_823'
    ]
    df_transaction['Portfolio_Name'] = np.select(conditions_port, choices_port, default='Other')

    # creating sec master table
    unique_cusips = pd.concat(
        [df_transaction['Security ID'], df_holding['CUSIP'], df_trader_info['Row Labels']]).dropna().drop_duplicates()
    df_sec_master = pd.DataFrame(unique_cusips, columns=['cusip'])
    df_sec_master = pd.merge(df_sec_master, df_trader_info, how='left', left_on='cusip', right_on='Row Labels')

    df_sec_master[['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating']] = df_sec_master[
        ['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating']].apply(lambda x: x.str.strip())
    df_sec_master[['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating']] = df_sec_master[
        ['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating']].fillna('NR')

    conditions_nrsro = [
        df_sec_master['SP Rating'] != 'NR',
        df_sec_master['Moody Rating'] != 'NR',
        df_sec_master['Fitch Rating'] != 'NR',
        df_sec_master['KBRA Rating'] != 'NR'
    ]
    choices_nrsro = [
        df_sec_master['SP Rating'],
        df_sec_master['Moody Rating'],
        df_sec_master['Fitch Rating'],
        df_sec_master['KBRA Rating']
    ]
    df_sec_master['NRSRO Rating'] = np.select(conditions_nrsro, choices_nrsro, default='NR')
    df_sec_master['NRSRO'] = np.select(conditions_nrsro, ['SP', 'Moody', 'Fitch', 'KBRA'], default='NR')

    # Create conditions, NEED TO CONFIRM THE MAPPING!!!!
    conditions_tranche = [
        df_sec_master['NRSRO Rating'].isin(['A', 'A-', 'A+', 'A2', 'A3', 'A1',]),
        df_sec_master['NRSRO Rating'].isin(['AA', 'AA2', 'AA1', 'AA-','Aa2', 'Aa1']),
        df_sec_master['NRSRO Rating'].isin(['B-', 'B1']),
        df_sec_master['NRSRO Rating'].isin(['BB-', 'BB+', 'BB', 'BA3', 'BA2', 'BA1','Ba3', 'Ba2', 'Ba1']),
        df_sec_master['NRSRO Rating'].isin(['BBB', 'BBB-', 'BBB+', 'BAA3', 'BAA2', 'Baa3', 'Baa2']),
        df_sec_master['NRSRO Rating'].isin(['CAA2','Caa2'])
    ]

    # Corresponding output values
    choices_tranche = ['A', 'AA', 'B', 'BB', 'BBB', 'CCC']

    # Apply the logic
    df_sec_master['Tranche Type'] = np.select(conditions_tranche, choices_tranche, default='NR')
    df_sec_master = df_sec_master.drop_duplicates()

    # creating positions table
    df_positions = df_holding[
        ['Portfolio Code', 'Investment Type', 'CUSIP',  'Current Face', 'BASEMarket Value','BASEOriginal Cost','Issuer Name',
         'Coupon Rate','Security Description', 'Maturity Date', 'Spread', 'Factor']]
    conditions_port = [
        df_positions['Portfolio Code'] == 'SBL_813_813',
        df_positions['Portfolio Code'] == 'SBL_814_814',
        df_positions['Portfolio Code'] == 'SBL_815_815',
        df_positions['Portfolio Code'] == 'SBL_816_816',
        df_positions['Portfolio Code'] == 'SBL_817_817',
        df_positions['Portfolio Code'] == 'SBL_821_821',
        df_positions['Portfolio Code'] == 'SBL_822_822',
        df_positions['Portfolio Code'] == 'SBL_823_823'
    ]
    df_positions['Portfolio_Name'] = np.select(conditions_port, choices_port, default='Other')
    df_positions = pd.merge(df_positions, df_sec_master[['Row Labels', 'Manager','WAL','Tranche Type']].drop_duplicates(), how='left', left_on='CUSIP',
                            right_on='Row Labels')

    # Create conditions
    df_positions['WAL'] = pd.to_numeric(df_positions['WAL'], errors='coerce')
    conditions_wal = [
        df_positions['WAL'] < 2,
        (df_positions['WAL'] >= 2) & (df_positions['WAL'] < 4),
        (df_positions['WAL'] >= 4) & (df_positions['WAL'] < 6),
        (df_positions['WAL'] >= 6) & (df_positions['WAL'] < 8),
        (df_positions['WAL'] >= 8) & (df_positions['WAL'] < 10),
        df_positions['WAL'] >= 10
    ]
    choices_wal = ['0<2', '2<4', '4<6', '6<8', '8<10', '10+']
    df_positions['WAL_Range'] = np.select(conditions_wal, choices_wal, default='unknown')

    df_positions['Coupon Rate'] = pd.to_numeric(df_positions['Coupon Rate'], errors='coerce')
    conditions_coupon = [
        df_positions['Coupon Rate'] < 3,
        (df_positions['Coupon Rate'] >= 3) & (df_positions['Coupon Rate'] < 4),
        (df_positions['Coupon Rate'] >= 4) & (df_positions['Coupon Rate'] < 5),
        (df_positions['Coupon Rate'] >= 5) & (df_positions['Coupon Rate'] < 6),
        (df_positions['Coupon Rate'] >= 6) & (df_positions['Coupon Rate'] < 7),
        df_positions['Coupon Rate'] >= 7
    ]
    choices_coupon= ['0<3', '3<4', '4<5', '5<6', '6<7', '7+']
    df_positions['coupon_Range'] = np.select(conditions_coupon, choices_coupon, default='0')

    return df_transaction, df_positions, df_sec_master



def update_report_tables(ws, table_data_dict, start_cell, gap, titles=None, formula_columns=None):
    """
    Update Excel tables in the worksheet with new data, optional section titles, and optional formulas.

    Args:
        ws: The worksheet object.
        table_data_dict: dict of {table_name: DataFrame}.
        start_cell: Cell string like 'B10' indicating where to begin writing.
        gap: Number of blank rows between tables.
        titles: Optional dict of {table_name: title string}.
        formula_columns: Optional dict of {table_name: {column_name: formula_string}}.
                        Formulas must use structured references if desired.
    """

    start_col_letter = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    start_col = ws[start_cell].column

    current_row = start_row

    for table_name, df in table_data_dict.items():
        if table_name not in ws.tables:
            raise ValueError(f"Table '{table_name}' not found in worksheet.")

        table = ws.tables[table_name]

        excel_columns = [col.name for col in table.tableColumns]

        total_cols = len(excel_columns)
        nrows = df.shape[0]

        col_name_to_index = {name: idx for idx, name in enumerate(excel_columns)}

        # Step 0: Write title if available
        if titles and table_name in titles:
            ws.cell(row=current_row, column=start_col, value=titles[table_name])
            current_row += 1

        # Step 1: Write header row
        for col_name, col_offset in col_name_to_index.items():
            ws.cell(row=current_row, column=start_col + col_offset, value=col_name)

        start_data_row = current_row + 1

        # Step 2: Write data rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_data_row):
            for df_col_idx, col_name in enumerate(df.columns):
                if col_name in col_name_to_index:
                    col_offset = col_name_to_index[col_name]
                    ws.cell(row=r_idx, column=start_col + col_offset, value=row[df_col_idx])


        # Step 3: Resize table range
        end_col = start_col + total_cols - 1
        end_row = start_data_row + nrows - 1
        table.ref = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{end_row}"

        # Step 4: Apply formulas if applicable
        if formula_columns and table_name in formula_columns:
            for col_name, formula_template  in formula_columns[table_name].items():
                if col_name not in col_name_to_index:
                    raise ValueError(f"Formula column '{col_name}' not found in table '{table_name}'")

                formula_col_offset = col_name_to_index[col_name]
                cusip_col_offset = col_name_to_index["Cusip"]
                print(f"[DEBUG] Applying formula to table '{table_name}', column '{col_name}'")
                print(f"[DEBUG] Formula content: {formula_template}")
                for r_idx in range(start_data_row, start_data_row + nrows):
                    cusip_cell = ws.cell(row=r_idx, column=start_col + cusip_col_offset).coordinate
                    # Replace placeholder in formula_template with actual cell ref
                    formula = formula_template.replace("{cusip_cell}", cusip_cell)
                    ws.cell(row=r_idx, column=start_col + formula_col_offset).value = formula

        # Debug print
        print(f"Updated table: {table_name}, ref: {table.ref}")


        # Step 5: Advance to next block
        current_row = end_row + gap + 1

    return


def generate_excel_report(date: datetime, inputdir: str, outputdir: str):
    formatted_date = date.strftime("%m-%d-%Y")
    # get all the raw data tables
    df_transaction, df_positions, df_sec_master = process_raw_data(formatted_date, inputdir)
    print('=========Processed all the raw data in :', inputdir, '=========')

    # get all the report table that doesn't need to be interactively refresh in the Excel
    df_manager = df_sec_master[['Manager']].dropna().drop_duplicates()
    df_manager.columns = ['CLO/ABS Debt']
    print('df_manager ', df_manager.shape)

    df_transaction['Type'] = np.select([df_transaction['Tran Type'] == 'MBS PMT', df_transaction['Tran Type'] == 'CALL',
                                            df_transaction['Tran Type'] == 'SELL', df_transaction['Tran Type'] == 'BUY'],
                                   ['Paydown', 'Call', 'SELL', 'BUY'], default='Other')
    df_transaction.drop(columns=['Tran Type'],inplace=True)

    df_paydown = df_transaction[df_transaction['Type'].isin(['Paydown', 'Call'])]

    df_paydown = (df_paydown[['Type', 'Security ID','Security Description','Settle Date', 'Coupon Rate','Quantity']]
        .drop_duplicates()
        .groupby(['Type', 'Security ID',
                  'Security Description'],
                 as_index=False)
        .agg({
            'Settle Date': list,
            'Coupon Rate': list,
            'Quantity': list
        })
    )

    # Compute weighted average of Coupon Rate
    df_paydown['Coupon Rate'] = df_paydown.apply(
        lambda row: sum(c * q for c, q in zip(row['Coupon Rate'],row['Quantity'])) / sum(row['Quantity'])
        if sum(row['Quantity']) != 0 else None,
        axis=1
    )
    df_paydown.drop( columns=['Quantity'],inplace=True)

    df_paydown = pd.merge(df_paydown,
                          df_sec_master[['cusip', 'Issuer Name', 'Tranche Type', 'NRSRO Rating']].drop_duplicates(),
                          how='left',
                          left_on='Security ID', right_on='cusip')

    df_paydown['Settle Date'] = df_paydown['Settle Date'].map(
        lambda dates: ', '.join(date.strftime('%Y-%m-%d') for date in dates))
    df_paydown = df_paydown.rename(columns={'Security ID': 'Cusip', 'Issuer Name': 'Issuer', 'Coupon Rate': 'Coupon'})
    print('df_paydown ', df_paydown.shape)

    df_sale = df_transaction[df_transaction['Type'].isin(['SELL'])]
    df_sale = df_sale[
        ['Security ID', 'Security Description', 'Settle Date', 'Trade Date', 'Maturity Date', 'Coupon Rate', 'Price']]
    df_sale = pd.merge(df_sale,
                       df_sec_master[['cusip', 'Issuer Name', 'Tranche Type', 'Market Price']].drop_duplicates(),
                       how='left', left_on='Security ID', right_on='cusip')
    df_sale = df_sale.rename(
        columns={'Security ID': 'Cusip', 'Issuer Name': 'Issuer', 'Coupon Rate': 'Coupon', 'Price': 'Sale Price',
                  'Tranche Type': 'Rating'})

    df_purchase = df_transaction[df_transaction['Type'].isin(['BUY'])]
    df_purchase = df_purchase[
        ['Security ID', 'Security Description', 'Settle Date', 'Trade Date', 'Maturity Date', 'Coupon Rate', 'Price']]
    df_purchase = pd.merge(df_purchase,
                       df_sec_master[['cusip', 'Issuer Name', 'Tranche Type', 'Market Price']].drop_duplicates(),
                       how='left', left_on='Security ID', right_on='cusip')
    df_purchase = df_purchase.rename(
        columns={'Security ID': 'Cusip', 'Issuer Name': 'Issuer', 'Coupon Rate': 'Coupon', 'Price': 'Purchase Price',
                 'Tranche Type': 'Rating'})

    print('=========Processed all the report table=========')

    df_holdings_report =  df_positions[['CUSIP','Issuer Name','Coupon Rate','Security Description', 'Issuer Name', 'Maturity Date', 'Spread', 'Factor']].drop_duplicates()
    df_holdings_report = pd.merge(df_holdings_report,
                       df_sec_master[['cusip', 'Tranche Type','Issue Date','NRSRO Rating','NRSRO','WAL','Par Sub','Implied DM','Next Payment Date','Non-Call Date']].drop_duplicates(),
                       how='left', left_on='CUSIP', right_on='cusip')
    df_holdings_report = df_holdings_report.rename(
        columns={'CUSIP': 'Cusip', 'Issuer Name': 'Issuer', 'Coupon Rate': 'Coupon'})
    df_holdings_report['Current Face'] = 0
    df_holdings_report['Market Value'] = 0
    df_A = df_holdings_report[df_holdings_report['Tranche Type'] == 'A']
    df_AA = df_holdings_report[df_holdings_report['Tranche Type'] == 'AA']
    df_B = df_holdings_report[df_holdings_report['Tranche Type'] == 'B']
    df_BB = df_holdings_report[df_holdings_report['Tranche Type'] == 'BB']
    df_BBB = df_holdings_report[df_holdings_report['Tranche Type'] == 'BBB']
    df_CCC = df_holdings_report[df_holdings_report['Tranche Type'] == 'CCC']


    #Row Labels	WAL	Next Payment Date 	Call Date
    print('=========Processed all the holdings table=========')
    print('=========Start to write Excel Report=========')

    base_dir = Path(__file__).parent.parent
    source_file = base_dir / 'report_template' / 'SkyRidge_CLO_ABS_Template.xlsx'
    source_file = str(source_file)
    # Write the DataFrames to an Excel file
    report_path = outputdir + "/SkyRidge_CLO_ABS_" + date.strftime("%Y%m%d") + ".xlsx"
    # Make a full file copy
    shutil.copy(source_file, report_path)
    with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_sec_master.to_excel(writer, sheet_name='Sec_Master', index=False)

    # Use openpyxl to populate tables
    wb = openpyxl.load_workbook(report_path)
    # Define each raw table and save to wb
    for df, sheet_name, table_name in [
        (df_positions, 'Positions', 'Table12'),
        (df_transaction, 'Transactions', 'Table16')
    ]:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        # Write new data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        max_row = r_idx  # from last iteration
        max_col = df.shape[1]
        end_col_letter = get_column_letter(max_col)
        table_range = f"A1:{end_col_letter}{max_row}"
        ws.tables[table_name].ref = table_range

    ws = wb['Report']
    table_data_dict = {
        'tbl_report_manager': df_manager,
        'tbl_report_paydown': df_paydown,
        'tbl_report_sale': df_sale,
        'tbl_report_purchase': df_purchase,
        'tbl_hold_a': df_A,
        'tbl_hold_aa': df_AA,
        'tbl_hold_b': df_B,
        'tbl_hold_bb': df_BB,
        'tbl_hold_bbb': df_BBB,
        'tbl_hold_ccc': df_CCC
    }
    titles = {
        'tbl_report_paydown': "Maturity/Paydown/Redemption/Call Report",
        'tbl_report_sale': "Sale Transactions",
        'tbl_report_purchase': "Purchase Transaction Report",
        'tbl_hold_a': "CLO/ABS Debt - A",
        'tbl_hold_aa': "CLO/ABS Debt - AA",
        'tbl_hold_b': "CLO/ABS Debt - B",
        'tbl_hold_bb': "CLO/ABS Debt - BB",
        'tbl_hold_bbb': "CLO/ABS Debt - BBB",
        'tbl_hold_ccc': "CLO/ABS Debt - CCC"
    }
    holdings_current_face = '=IF($C$4="All Portfolios",SUMIFS(Table12[Current Face],Table12[CUSIP],{cusip_cell}),SUMIFS(Table12[Current Face],Table12[Portfolio_Name],$C$4,Table12[CUSIP],{cusip_cell}))'
    holdings_market_value = '=IF($C$4="All Portfolios",SUMIFS(Table12[BASEMarket Value],Table12[CUSIP],{cusip_cell}),SUMIFS(Table12[BASEMarket Value],Table12[Portfolio_Name],$C$4,Table12[CUSIP],{cusip_cell}))'

    formula_columns = {
        'tbl_report_paydown': {
            'Face Value': '=IF($C$4="All Portfolios",-SUMIFS(Table16[Quantity],Table16[Security ID],{cusip_cell},Table16[Type], "Paydown"),-SUMIFS(Table16[Quantity],Table16[Portfolio_Name],$C$4,Table16[Security ID],{cusip_cell},Table16[Type], "Paydown"))',
        },
        'tbl_report_sale': {
            'Face Value': '=IF($C$4="All Portfolios",SUMIFS(Table16[Quantity],Table16[Security ID],{cusip_cell},Table16[Type], "SELL"),SUMIFS(Table16[Quantity],Table16[Portfolio_Name],$C$4,Table16[Security ID],{cusip_cell},Table16[Type], "SELL"))',
            'Market Value': '=IF($C$4="All Portfolios",SUMIFS(Table16[Cost Proceeds],Table16[Security ID],{cusip_cell},Table16[Type], "SELL"),SUMIFS(Table16[Cost Proceeds],Table16[Portfolio_Name],$C$4,Table16[Security ID],{cusip_cell},Table16[Type], "SELL"))'
        },
        'tbl_report_purchase': {
            'Face Value': '=IF($C$4="All Portfolios",SUMIFS(Table16[Quantity],Table16[Security ID],{cusip_cell},Table16[Type], "BUY"),SUMIFS(Table16[Quantity],Table16[Portfolio_Name],$C$4,Table16[Security ID],{cusip_cell},Table16[Type], "BUY"))',
            'Market Value': '=IF($C$4="All Portfolios",SUMIFS(Table16[Cost Proceeds],Table16[Security ID],{cusip_cell},Table16[Type], "BUY"),SUMIFS(Table16[Cost Proceeds],Table16[Portfolio_Name],$C$4,Table16[Security ID],{cusip_cell},Table16[Type], "BUY"))'
        },

        'tbl_hold_a': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_aa': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_b': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_bb': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_bbb': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_ccc': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        }
    }
    update_report_tables(ws, table_data_dict, start_cell='B35', gap=5,titles=titles, formula_columns= formula_columns)

    wb.save(report_path)
    print('=========Saved to directory:', report_path, '=========')

    return


"""
    # Use openpyxl to populate tables
    wb = openpyxl.load_workbook(report_path)
    # Define each raw table and save to wb
    for df, sheet_name, table_name in [
        (df_transaction, 'Transactions', 'tbl_transactions'),
        (df_positions, 'Positions', 'tbl_positions'),
        (df_sec_master, 'Sec_Master', 'tbl_sec_master')
    ]:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        # Write new data including headers
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
                
                
    ========
    with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_transaction.to_excel(writer, sheet_name='Transactions', index=False)
        df_positions.to_excel(writer, sheet_name='Positions', index=False)
        df_sec_master.to_excel(writer, sheet_name='Sec_Master', index=False)

"""
