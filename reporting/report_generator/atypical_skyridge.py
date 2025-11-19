import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime, date
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from reporting.utils import update_report_tables, get_tranche_df, get_empty_df


# palceholder for datasets, to be replaced with Snowflake holdings table, transaction table and security master table
#  . .\.venv\Scripts\Activate.ps1
# dir = "C:/Users/XiaHanlu/workspace/legacyReporting/local_reading/panagram"
# date = '04-30-2025'
def process_raw_data(inputdir: str, choices_port: list, target_portfolio_list: list, holdings_file : str, transaction_file: str, input_from_jared :str):
    map_dict = dict(zip(target_portfolio_list, choices_port))
    # read in the holdings table
    df_holding = pd.read_excel(inputdir + holdings_file,
                               sheet_name='DBO_PANAGRAM_SBLHoldingFile_NoF', index_col=None, header=0)
    # nake sure there is no space in column names
    df_holding.columns = df_holding.columns.str.strip()
    # filter only for portfolios that we care
    df_holding = df_holding[df_holding['Portfolio Code'].isin(target_portfolio_list)]  #
    # filter out all the positions that already have MV = 0
    df_holding = df_holding[df_holding['BASEMarket Value'] != 0]

    # read in the data we requested from the investment team, contact John Rozario and Aaron Zhang for details
    df_trader_info =  pd.read_csv(inputdir + input_from_jared)
    df_trader_info['Manager'] = df_trader_info['Manager'].fillna('Unknown')

    # creating transactions table
    df_transaction = pd.read_excel(inputdir + transaction_file, sheet_name='DBO_Security Transactions', index_col=None, header=0,
                                   usecols=['Portfolio', 'Security ID', 'Tran Type', 'Security Description', 'Coupon Rate',
                                            'Trade Date', 'Settle Date', 'Maturity Date', 'Quantity', 'Price', 'Cost Proceeds'])
    # filter only for portfolios that we care
    df_transaction = df_transaction[df_transaction['Portfolio'].isin(target_portfolio_list)]
    df_transaction['Portfolio_Name'] = df_transaction['Portfolio'].map(map_dict).fillna('Other')

    # creating sec master table
    unique_cusips = pd.concat(
        [df_transaction['Security ID'], df_holding['CUSIP'], df_trader_info['Row Labels']]).dropna().drop_duplicates()
    df_sec_master = pd.DataFrame(unique_cusips, columns=['cusip'])
    df_sec_master = pd.merge(df_sec_master, df_trader_info, how='left', left_on='cusip', right_on='Row Labels')

    #df_sec_master[['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating', 'DBRS Rating']] = df_sec_master[['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating', 'DBRS Rating']].astype(str).apply(lambda x: x.str.strip())
    df_sec_master[['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating', 'DBRS Rating']] = df_sec_master[
        ['SP Rating', 'Moody Rating', 'Fitch Rating', 'KBRA Rating', 'DBRS Rating']].fillna('NR')

    conditions_nrsro = [
        df_sec_master['SP Rating'] != 'NR',
        df_sec_master['Moody Rating'] != 'NR',
        df_sec_master['Fitch Rating'] != 'NR',
        df_sec_master['KBRA Rating'] != 'NR',
        df_sec_master['DBRS Rating'] != 'NR'
    ]
    choices_nrsro = [
        df_sec_master['SP Rating'],
        df_sec_master['Moody Rating'],
        df_sec_master['Fitch Rating'],
        df_sec_master['KBRA Rating'],
        df_sec_master['DBRS Rating']
    ]
    df_sec_master['NRSRO Rating'] = np.select(conditions_nrsro, choices_nrsro, default='NR')
    df_sec_master['NRSRO'] = np.select(conditions_nrsro, ['SP', "Moody's", 'Fitch', 'KBRA', 'DBRS'], default='NR')

    # Create conditions, NEED TO CONFIRM THE MAPPING!!!!
    conditions_tranche = [
        df_sec_master['NRSRO Rating'].isin(['A', 'A-', 'A+', 'A2', 'A3', 'A1', ]),
        df_sec_master['NRSRO Rating'].isin(['AA', 'AA+', 'AA-', 'AA3', 'AA2', 'AA1', 'Aa3', 'Aa2', 'Aa1', 'AAL']),
        df_sec_master['NRSRO Rating'].isin(['AAA', 'Aaa']),
        df_sec_master['NRSRO Rating'].isin(['B', 'B+', 'B-', 'B1', 'B2', 'B3']),
        df_sec_master['NRSRO Rating'].isin(['BB-', 'BB+', 'BB', 'BA3', 'BA2', 'BA1', 'Ba3', 'Ba2', 'Ba1']),
        df_sec_master['NRSRO Rating'].isin(['BBB', 'BBB-', 'BBB+', 'BAA3', 'BAA2', 'BAA1', 'Baa3', 'Baa2', 'Baa1']),
        df_sec_master['NRSRO Rating'].isin(['CCC', 'CCC-', 'CCC+', 'CAA1', 'Caa1', 'CAA2', 'Caa2', 'CAA3', 'Caa3']),
        df_sec_master['NRSRO Rating'].isin(['NR', 'WR', 'D'])
    ]

    # Corresponding output values
    choices_tranche = ['A', 'AA', 'AAA', 'B', 'BB', 'BBB', 'CCC', 'NR']

    # Apply the logic for tranche rating
    df_sec_master['Tranche Rating'] = np.select(conditions_tranche, choices_tranche, default='NR')
    df_sec_master = df_sec_master.drop_duplicates()

    # creating positions table
    df_positions = df_holding[
        ['Portfolio Code','Investment Type', 'CUSIP', 'Current Face', 'BASEMarket Value', 'BASEOriginal Cost', 'Issuer Name',
         'Coupon Rate', 'Security Description', 'Maturity Date', 'Spread', 'Factor']]


    df_positions['Portfolio_Name'] = df_positions['Portfolio Code'].map(map_dict).fillna('Other')
    df_positions = pd.merge(df_positions, df_sec_master[['Row Labels', 'Manager', 'WAL', 'Tranche Rating']].drop_duplicates(), how='left',
                            left_on='CUSIP',
                            right_on='Row Labels')

    # Create conditions
    df_positions['WAL'] = pd.to_numeric(df_positions['WAL'], errors='coerce')
    conditions_wal = [
        df_positions['WAL'] < 2,
        (df_positions['WAL'] >= 2) & (df_positions['WAL'] < 4),
        (df_positions['WAL'] >= 4) & (df_positions['WAL'] < 6),
        (df_positions['WAL'] >= 6) & (df_positions['WAL'] < 8),
        (df_positions['WAL'] >= 8) & (df_positions['WAL'] < 10),
        df_positions['WAL'] >= 10
    ]
    choices_wal = ['0<2', '2<4', '4<6', '6<8', '8<10', '10+']
    df_positions['WAL_Range'] = np.select(conditions_wal, choices_wal, default='unknown')

    df_positions['Coupon Rate'] = pd.to_numeric(df_positions['Coupon Rate'], errors='coerce')
    conditions_coupon = [
        df_positions['Coupon Rate'] < 3,
        (df_positions['Coupon Rate'] >= 3) & (df_positions['Coupon Rate'] < 4),
        (df_positions['Coupon Rate'] >= 4) & (df_positions['Coupon Rate'] < 5),
        (df_positions['Coupon Rate'] >= 5) & (df_positions['Coupon Rate'] < 6),
        (df_positions['Coupon Rate'] >= 6) & (df_positions['Coupon Rate'] < 7),
        df_positions['Coupon Rate'] >= 7
    ]
    choices_coupon = ['0<3', '3<4', '4<5', '5<6', '6<7', '7+']
    df_positions['coupon_Range'] = np.select(conditions_coupon, choices_coupon, default='0')

    return df_transaction, df_positions, df_sec_master


def generate_excel_report(date: str, inputdir: str, outputdir: str):
    # ============================configuration section FOR SkyRidge Report=========================

    #"""
    choices_port = ['44G8/813/IG CLO FWH', '44G9/814/ABS FWH', '44H1/815/Atypical FWH', '44I4/816/HY CLOs FWH', 'P44I5/817/HY CLOs OC', '44J6/821/IG CLO OC',
                   '44J7/822/ABS OC', '44J5/823/Atypical OC']
    target_portfolio_list = ['SBL_813_813','SBL_814_814','SBL_815_815','SBL_816_816','SBL_817_817','SBL_821_821','SBL_822_822','SBL_823_823']
    output_filename = 'SkyRidge_CLO_ABS'
    template_filename = 'SkyRidge_CLO_ABS_Template.xlsx'
    holdings_file = 'Panagram Holding File 10-31-2025.xlsx'
    transaction_file = 'Security Transactions_sky20251031.xlsx'
    input_from_jared = 'Data_request_20251031.csv'
    manager_table_start = 'B38'
    #"""

    # ============================configuration section End=========================
    formatted_date = date.strftime("%m-%d-%Y")
    # get all the raw data tables
    df_transaction, df_positions, df_sec_master = process_raw_data(inputdir, choices_port, target_portfolio_list, holdings_file, transaction_file, input_from_jared)
    print('=========Processed all the raw data in :', inputdir, '=========')

    # Prepare the Manager Table
    df_manager = df_sec_master[['Manager']].dropna().drop_duplicates()
    df_manager.columns = ['Manager Name']
    print('df_manager ', df_manager.shape)

    # Prepare the transaction table
    df_transaction['Type'] = np.select([df_transaction['Tran Type'] == 'MBS PMT', df_transaction['Tran Type'] == 'CALL',
                                        df_transaction['Tran Type'] == 'SELL', df_transaction['Tran Type'] == 'BUY'],
                                       ['Paydown', 'Call', 'SELL', 'BUY'], default='Other')
    df_transaction.drop(columns=['Tran Type'], inplace=True)
    df_transaction['Maturity Date'] = pd.to_datetime(df_transaction['Maturity Date']).dt.strftime("%m-%d-%Y")
    df_transaction['Trade Date'] = pd.to_datetime(df_transaction['Trade Date']).dt.strftime("%m-%d-%Y")
    df_transaction['Settle Date'] = pd.to_datetime(df_transaction['Settle Date']).dt.strftime("%m-%d-%Y")

    # Prepare the paydown table
    df_paydown = df_transaction[df_transaction['Type'].isin(['Paydown', 'Call'])]
    df_paydown = (df_paydown[['Type', 'Security ID', 'Security Description', 'Settle Date']]
    .drop_duplicates()
    .groupby(['Type', 'Security ID', 'Security Description'],
             as_index=False)
    .agg({
        'Settle Date': list
    })
    )
    df_paydown['Settle Date'] = df_paydown['Settle Date'].map(lambda dates: ', '.join(dates))

    # Compute weighted average of Coupon Rate
    """
        df_paydown['Coupon Rate'] = df_paydown.apply(
        lambda row: sum(c * q for c, q in zip(row['Coupon Rate'],row['Quantity'])) / sum(row['Quantity'])
        if sum(row['Quantity']) != 0 else None,axis=1
    )
    df_paydown.drop( columns=['Quantity'],inplace=True)
    """
    df_paydown = pd.merge(df_paydown,
                          df_sec_master[['cusip', 'Issuer Name', 'Tranche Rating']].drop_duplicates(), how='left',
                          left_on='Security ID', right_on='cusip')
    df_paydown = df_paydown.rename(columns={'Security ID': 'Cusip', 'Issuer Name': 'Issuer'})
    print('df_paydown ', df_paydown.shape)

    # Prepare the sales table
    df_sale = df_transaction[df_transaction['Type'].isin(['SELL'])]
    df_sale = (df_sale[['Security ID', 'Security Description', 'Settle Date', 'Trade Date', 'Maturity Date', 'Price']]
    .drop_duplicates()
    .groupby(['Security ID', 'Security Description', 'Trade Date', 'Maturity Date', 'Price'],
             as_index=False)
    .agg({
        'Settle Date': list
    })
    )
    df_sale['Settle Date'] = df_sale['Settle Date'].map(lambda dates: ', '.join(dates))
    df_sale = pd.merge(df_sale,
                       df_sec_master[['cusip', 'Issuer Name', 'Tranche Rating', 'Market Price']].drop_duplicates(),
                       how='left', left_on='Security ID', right_on='cusip')
    df_sale = df_sale.rename(
        columns={'Security ID': 'Cusip', 'Issuer Name': 'Issuer', 'Price': 'Sale Price'})

    # Prepare the purchase table
    df_purchase = df_transaction[df_transaction['Type'].isin(['BUY'])]
    df_purchase = (df_purchase[['Security ID', 'Security Description', 'Settle Date', 'Trade Date', 'Maturity Date', 'Price']]
    .drop_duplicates()
    .groupby(['Security ID', 'Security Description', 'Trade Date', 'Maturity Date', 'Price'],
             as_index=False)
    .agg({
        'Settle Date': list
    })
    )
    df_purchase['Settle Date'] = df_purchase['Settle Date'].map(lambda dates: ', '.join(dates))
    df_purchase = pd.merge(df_purchase,
                           df_sec_master[['cusip', 'Issuer Name', 'Tranche Rating', 'Market Price']].drop_duplicates(),
                           how='left', left_on='Security ID', right_on='cusip')
    df_purchase = df_purchase.rename(
        columns={'Security ID': 'Cusip', 'Issuer Name': 'Issuer', 'Price': 'Purchase Price'})

    df_sale = get_empty_df(df_sale)

    print('=========Processed all the report table=========')

    df_holdings_report = df_positions[
        ['CUSIP', 'Issuer Name', 'Coupon Rate', 'Security Description', 'Issuer Name', 'Maturity Date', 'Spread', 'Factor']].drop_duplicates()
    df_holdings_report = pd.merge(df_holdings_report,
                                  df_sec_master[['cusip', 'Tranche Rating', 'Issue Date', 'Market Price', 'NRSRO Rating', 'NRSRO', 'WAL', 'Par Sub',
                                                 'Implied DM', 'Next Payment Date', 'Non-Call Date']].drop_duplicates(),
                                  how='left', left_on='CUSIP', right_on='cusip')
    df_holdings_report = df_holdings_report.rename(
        columns={'CUSIP': 'Cusip', 'Issuer Name': 'Issuer', 'Coupon Rate': 'Coupon'})

    df_holdings_report['Issue Date'] = pd.to_datetime(df_holdings_report['Issue Date'], format='%Y%m%d').dt.strftime('%Y-%m-%d')
    df_holdings_report['Par Sub'] = df_holdings_report['Par Sub'] / 100
    df_holdings_report['Maturity Date'] = pd.to_datetime(df_holdings_report['Maturity Date']).dt.strftime("%m-%d-%Y")
    df_holdings_report['Next Payment Date'] = pd.to_datetime(df_holdings_report['Next Payment Date']).dt.strftime("%m-%d-%Y")
    df_holdings_report['Non-Call Date'] = pd.to_datetime(df_holdings_report['Non-Call Date']).dt.strftime("%m-%d-%Y")
    df_holdings_report['Current Face'] = 0
    df_holdings_report['Market Value'] = 0

    df_A = get_tranche_df(df_holdings_report, 'A')
    df_AA = get_tranche_df(df_holdings_report, 'AA')
    df_AAA = get_tranche_df(df_holdings_report, 'AAA')
    df_B = get_tranche_df(df_holdings_report, 'B')
    df_BB = get_tranche_df(df_holdings_report, 'BB')
    df_BBB = get_tranche_df(df_holdings_report, 'BBB')
    df_CCC = get_tranche_df(df_holdings_report, 'CCC')
    df_NR = get_tranche_df(df_holdings_report, 'NR')

    # Row Labels	WAL	Next Payment Date 	Call Date
    print('=========Processed all the holdings table=========')
    print('=========Start to write Excel Report=========')

    base_dir = Path(__file__).parent.parent
    source_file = base_dir / 'report_template' / template_filename
    source_file = str(source_file)
    # Write the DataFrames to an Excel file
    report_path = outputdir + "/" + output_filename + date.strftime("%Y%m%d") + ".xlsx"
    # Make a full file copy
    shutil.copy(source_file, report_path)
    with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_sec_master.to_excel(writer, sheet_name='Sec_Master', index=False)

    # Use openpyxl to populate tables
    wb = openpyxl.load_workbook(report_path)
    # Define each raw table and save to wb
    for df, sheet_name, table_name in [
        (df_positions, 'Positions', 'tbl_rawhold'),
        (df_transaction, 'Transactions', 'tbl_rawtran')
    ]:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        # Write new data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        max_row = r_idx  # from last iteration
        max_col = df.shape[1]
        end_col_letter = get_column_letter(max_col)
        table_range = f"A1:{end_col_letter}{max_row}"
        # ws.tables[table_name].ref = table_range

    ws = wb['Report']
    table_data_dict = {
        'tbl_report_manager': df_manager,
        'tbl_report_paydown': df_paydown,
        'tbl_report_sale': df_sale,
        'tbl_report_purchase': df_purchase,
        'tbl_hold_aaa': df_AAA,
        'tbl_hold_aa': df_AA,
        'tbl_hold_a': df_A,
        'tbl_hold_bbb': df_BBB,
        'tbl_hold_bb': df_BB,
        'tbl_hold_b': df_B,
        'tbl_hold_ccc': df_CCC,
        'tbl_hold_nr': df_NR
    }
    titles = {
        'tbl_report_paydown': "Maturity/Paydown/Redemption/Call Report",
        'tbl_report_sale': "Sale Transactions",
        'tbl_report_purchase': "Purchase Transaction Report",
        'tbl_hold_aaa': "CLO/ABS Debt - AAA",
        'tbl_hold_aa': "CLO/ABS Debt - AA",
        'tbl_hold_a': "CLO/ABS Debt - A",
        'tbl_hold_bbb': "CLO/ABS Debt - BBB",
        'tbl_hold_bb': "CLO/ABS Debt - BB",
        'tbl_hold_b': "CLO/ABS Debt - B",
        'tbl_hold_ccc': "CLO/ABS Debt - CCC",
        'tbl_hold_nr': "CLO/ABS Debt - NR"
    }
    holdings_current_face = '=IF($C$4="All Portfolios",SUMIFS(tbl_rawhold[Current Face],tbl_rawhold[CUSIP],{cusip_cell}),SUMIFS(tbl_rawhold[Current Face],tbl_rawhold[Portfolio_Name],$C$4,tbl_rawhold[CUSIP],{cusip_cell}))'
    holdings_market_value = '=IF($C$4="All Portfolios",SUMIFS(tbl_rawhold[BASEMarket Value],tbl_rawhold[CUSIP],{cusip_cell}),SUMIFS(tbl_rawhold[BASEMarket Value],tbl_rawhold[Portfolio_Name],$C$4,tbl_rawhold[CUSIP],{cusip_cell}))'

    formula_columns = {
        'tbl_report_paydown': {
            'Face Value': '=IF($C$4="All Portfolios",-SUMIFS(tbl_rawtran[Quantity],tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "Paydown"),-SUMIFS(tbl_rawtran[Quantity],tbl_rawtran[Portfolio_Name],$C$4,tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "Paydown"))',
        },
        'tbl_report_sale': {
            'Face Value': '=IF($C$4="All Portfolios",SUMIFS(tbl_rawtran[Quantity],tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "SELL"),SUMIFS(tbl_rawtran[Quantity],tbl_rawtran[Portfolio_Name],$C$4,tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "SELL"))'
            # 'Market Value': '=IF($C$4="All Portfolios",SUMIFS(tbl_rawtran[Cost Proceeds],tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "SELL"),SUMIFS(tbl_rawtran[Cost Proceeds],tbl_rawtran[Portfolio_Name],$C$4,tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "SELL"))'
        },
        'tbl_report_purchase': {
            'Face Value': '=IF($C$4="All Portfolios",SUMIFS(tbl_rawtran[Quantity],tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "BUY"),SUMIFS(tbl_rawtran[Quantity],tbl_rawtran[Portfolio_Name],$C$4,tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "BUY"))'
            # 'Market Value': '=IF($C$4="All Portfolios",SUMIFS(tbl_rawtran[Cost Proceeds],tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "BUY"),SUMIFS(tbl_rawtran[Cost Proceeds],tbl_rawtran[Portfolio_Name],$C$4,tbl_rawtran[Security ID],{cusip_cell},tbl_rawtran[Type], "BUY"))'
        },
        'tbl_hold_aaa': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_aa': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_a': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_bbb': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_bb': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_b': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_ccc': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        },
        'tbl_hold_nr': {
            'Current Face': holdings_current_face,
            'Market Value': holdings_market_value
        }
    }
    update_report_tables(ws, table_data_dict, start_cell=manager_table_start, gap=5, titles=titles, formula_columns=formula_columns)

    wb.save(report_path)
    print('=========Saved to directory:', report_path, '=========')

    return


"""
    # Use openpyxl to populate tables
    wb = openpyxl.load_workbook(report_path)
    # Define each raw table and save to wb
    for df, sheet_name, table_name in [
        (df_transaction, 'Transactions', 'tbl_transactions'),
        (df_positions, 'Positions', 'tbl_positions'),
        (df_sec_master, 'Sec_Master', 'tbl_sec_master')
    ]:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        # Write new data including headers
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)


    ========
    with pd.ExcelWriter(report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_transaction.to_excel(writer, sheet_name='Transactions', index=False)
        df_positions.to_excel(writer, sheet_name='Positions', index=False)
        df_sec_master.to_excel(writer, sheet_name='Sec_Master', index=False)

"""
